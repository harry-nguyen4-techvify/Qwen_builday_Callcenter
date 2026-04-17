"""
Interactive flow designer: user picks cell pairs from an Excel file.

Usage:
    uv run python examples/design_flow.py
    uv run python examples/design_flow.py --xlsx path/to/form.xlsx
    uv run python examples/design_flow.py --refine intake_form --feedback "Thêm node summary"

Workflow:
    1. Load Excel → hiển thị nội dung các ô
    2. User chọn từng cặp: ô nhãn (label) + ô cần điền (input)
    3. Gọi FlowDesigner (Qwen API) → sinh FlowModel
    4. Lưu vào flows/ hoặc in preview
"""

from __future__ import annotations

import argparse
import asyncio
import pathlib
import re
import sys
import time

sys.path.insert(0, str(pathlib.Path(__file__).parent.parent))

from dotenv import load_dotenv
load_dotenv()


# ── Excel display ──────────────────────────────────────────────────────────────

def _col_letter(col: int) -> str:
    """Convert 1-based column index to letter (1→A, 26→Z, 27→AA)."""
    result = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def _parse_cell_ref(ref: str, sheet_name: str) -> tuple[str, str]:
    """
    Parse user input like 'B3', 'IntakeForm!B3', 'b3' → (sheet, coord).
    Returns (sheet_name, 'B3') — coord always upper-cased.
    """
    ref = ref.strip()
    if "!" in ref:
        sheet, coord = ref.split("!", 1)
        return sheet.strip(), coord.strip().upper()
    return sheet_name, ref.upper()


def load_excel_cells(xlsx: pathlib.Path) -> dict[str, dict[str, object]]:
    """
    Returns { sheet_name: { 'B3': value, ... } } for all non-empty cells.
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    result: dict[str, dict[str, object]] = {}
    for ws in wb.worksheets:
        cells: dict[str, object] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cells[cell.coordinate] = cell.value
        result[ws.title] = cells
    return result


def print_excel_table(xlsx: pathlib.Path) -> None:
    """Print every non-empty cell in a readable table per sheet."""
    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx), data_only=True)

    for ws in wb.worksheets:
        print(f"\n  Sheet: {ws.title}")
        print(f"  {'Cell':<8} {'Value'}")
        print(f"  {'-'*8} {'-'*45}")
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value)[:45]
                    print(f"  {cell.coordinate:<8} {val}")


# ── Interactive field picker ───────────────────────────────────────────────────

def _slugify(text: str) -> str:
    text = text.lower().strip().rstrip(":")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_") or "field"


def pick_fields_interactively(
    xlsx: pathlib.Path,
) -> tuple[list, str, str]:
    """
    Show Excel content, let user pick (label_cell, input_cell) pairs.
    Returns (field_definitions, raw_text, sheet_name).
    """
    from core.models.field_defs import FieldDefinition, FieldType, FieldConstraints

    print_excel_table(xlsx)

    import openpyxl
    wb = openpyxl.load_workbook(str(xlsx), data_only=True)
    sheet_names = [ws.title for ws in wb.worksheets]

    # Pick sheet if multiple
    default_sheet = sheet_names[0]
    if len(sheet_names) > 1:
        print(f"\n  Sheets: {', '.join(sheet_names)}")
        chosen = input(f"  Chọn sheet [{default_sheet}]: ").strip()
        default_sheet = chosen if chosen in sheet_names else default_sheet

    ws = wb[default_sheet]
    cell_map: dict[str, str] = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_map[cell.coordinate] = str(cell.value)

    # Build raw_text for context
    raw_text = "\n".join(
        f"{default_sheet}!{coord}: {val}"
        for coord, val in cell_map.items()
    )

    print(f"\n  Nhập từng cặp ô. Ví dụ: ô nhãn = A3, ô cần điền = B3")
    print(f"  Gõ 'done' hoặc để trống khi xong.\n")

    field_defs: list[FieldDefinition] = []
    seen_ids: dict[str, int] = {}
    idx = 1

    while True:
        # --- label cell ---
        raw_label = input(f"  [{idx}] Ô nhãn    (ví dụ A3, hoặc Enter để kết thúc): ").strip()
        if not raw_label or raw_label.lower() == "done":
            break

        _, label_coord = _parse_cell_ref(raw_label, default_sheet)
        label_val = cell_map.get(label_coord)
        if label_val is None:
            print(f"      Ô {label_coord} trống hoặc không tồn tại. Thử lại.")
            continue
        print(f"      → Nhãn: \"{label_val}\"")

        # --- input cell ---
        raw_input = input(f"  [{idx}] Ô cần điền (ví dụ B3): ").strip()
        if not raw_input or raw_input.lower() == "done":
            print("      Bỏ qua trường này.")
            continue

        _, input_coord = _parse_cell_ref(raw_input, default_sheet)

        # --- field id ---
        suggested_id = _slugify(str(label_val))
        base_id = suggested_id
        if base_id in seen_ids:
            seen_ids[base_id] += 1
            suggested_id = f"{base_id}_{seen_ids[base_id]}"
        else:
            seen_ids[base_id] = 0

        raw_id = input(f"  [{idx}] Field ID   [{suggested_id}]: ").strip()
        field_id = raw_id if raw_id else suggested_id

        fd = FieldDefinition(
            id=field_id,
            label=str(label_val).rstrip(":"),
            cell_ref=f"{default_sheet}!{input_coord}",
            type=FieldType.TEXT,
            constraints=FieldConstraints(),
        )
        field_defs.append(fd)
        print(f"      ✓ {field_id:30s} → {default_sheet}!{input_coord}\n")
        idx += 1

    return field_defs, raw_text, default_sheet


# ── Design command ─────────────────────────────────────────────────────────────

async def cmd_design(xlsx: pathlib.Path, prompt: str, save: bool) -> None:
    from core.designer.flow_designer import FlowDesigner
    from core.store.flow_store import FlowStore

    print(f"\n=== FLOW DESIGNER ===")
    print(f"File: {xlsx}\n")

    # Step 1: interactive field picking
    field_defs, raw_text, _ = pick_fields_interactively(xlsx)

    if not field_defs:
        print("\nKhông có field nào được chọn. Thoát.")
        return

    print(f"\n  Đã chọn {len(field_defs)} field(s):")
    for fd in field_defs:
        print(f"    • {fd.id:30s} {fd.cell_ref}")

    # Step 2: prompt
    if not prompt:
        custom = input("\n  Mô tả flow (Enter để dùng mặc định): ").strip()
        prompt = custom or "Collect form data from user via voice agent in Vietnamese"
    print(f"  Prompt: {prompt!r}")

    # Step 3: call FlowDesigner
    print(f"\n  Đang gọi Qwen API…")
    designer = FlowDesigner()
    t0 = time.perf_counter()
    flow = await designer.design(
        field_defs=field_defs,
        user_prompt=prompt,
        template_raw=raw_text,
    )
    elapsed = time.perf_counter() - t0

    print(f"  Xong ({elapsed:.1f}s)")
    print(f"\n  flow_id : {flow.flow_id}")
    print(f"  name    : {flow.name}")
    print(f"  nodes   : {len(flow.nodes)}")
    print(f"  edges   : {len(flow.edges)}")
    node_types: dict[str, int] = {}
    for n in flow.nodes:
        node_types[n.type] = node_types.get(n.type, 0) + 1
    for t, c in sorted(node_types.items()):
        print(f"    {t:20s} x{c}")

    if save:
        path = FlowStore().save(flow)
        print(f"\n  Saved → {path}")
        print(f"  Chạy agent với metadata: {{\"flow_id\": \"{flow.flow_id}\"}}")
    else:
        ans = input("\n  Lưu flow vào flows/? [y/N]: ").strip().lower()
        if ans == "y":
            path = FlowStore().save(flow)
            print(f"  Saved → {path}")
            print(f"  Chạy agent với metadata: {{\"flow_id\": \"{flow.flow_id}\"}}")
        else:
            print("\n--- JSON preview (60 dòng đầu) ---")
            lines = flow.model_dump_json(indent=2).splitlines()
            print("\n".join(lines[:60]))
            if len(lines) > 60:
                print(f"  … ({len(lines) - 60} dòng nữa)")


# ── Refine command ─────────────────────────────────────────────────────────────

async def cmd_refine(flow_id: str, feedback: str, save: bool) -> None:
    from core.designer.flow_designer import FlowDesigner
    from core.store.flow_store import FlowStore

    store = FlowStore()
    if not store.exists(flow_id):
        print(f"ERROR: flow '{flow_id}' không tìm thấy trong flows/")
        sys.exit(1)

    current = store.load(flow_id)
    print(f"\nFlow '{flow_id}' — version {current.version}, {len(current.nodes)} nodes")

    if not feedback:
        feedback = input("Yêu cầu chỉnh sửa: ").strip()
    if not feedback:
        print("Không có feedback. Thoát.")
        return

    print(f"Đang refine…")
    t0 = time.perf_counter()
    refined = await FlowDesigner().refine(current, feedback)
    print(f"Xong ({time.perf_counter() - t0:.1f}s) — version {refined.version}, {len(refined.nodes)} nodes")

    if save:
        path = store.save(refined)
        print(f"Saved → {path}")
    else:
        ans = input("Lưu? [y/N]: ").strip().lower()
        if ans == "y":
            store.save(refined)
            print("Saved.")
        else:
            lines = refined.model_dump_json(indent=2).splitlines()
            print("\n".join(lines[:60]))


# ── Entry point ────────────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(description="Interactive FlowDesigner")
    sub = parser.add_subparsers(dest="cmd")

    p_d = sub.add_parser("design", help="Tạo flow mới từ Excel")
    p_d.add_argument("--xlsx", default="examples/intake_form.xlsx")
    p_d.add_argument("--prompt", default="")
    p_d.add_argument("--save", action="store_true")

    p_r = sub.add_parser("refine", help="Chỉnh sửa flow đã lưu")
    p_r.add_argument("flow_id")
    p_r.add_argument("--feedback", default="")
    p_r.add_argument("--save", action="store_true")

    args = parser.parse_args()
    if args.cmd is None:
        args.cmd = "design"
        args.xlsx = "examples/intake_form.xlsx"
        args.prompt = ""
        args.save = False

    if args.cmd == "design":
        asyncio.run(cmd_design(pathlib.Path(args.xlsx), args.prompt, args.save))
    elif args.cmd == "refine":
        asyncio.run(cmd_refine(args.flow_id, args.feedback, args.save))


if __name__ == "__main__":
    main()
