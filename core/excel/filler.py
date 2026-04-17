from __future__ import annotations

import re
from pathlib import Path

import openpyxl

from core.models.form_data import FormData


# Regex matching an optional sheet qualifier: "Sheet1!B3" or just "B3".
_SHEET_CELL_RE = re.compile(r"^(?:(.+)!)?([A-Z]+[0-9]+)$", re.IGNORECASE)


class ExcelFiller:
    """Write FormData values into an xlsx template and save to output_path."""

    def fill(
        self,
        template_path: str | Path,
        form_data: FormData,
        cell_mapping: dict[str, str],
        output_path: str | Path,
    ) -> Path:
        """
        Parameters
        ----------
        template_path : path to the source xlsx template.
        form_data     : FormData instance holding collected field values.
        cell_mapping  : mapping of field_id -> cell_ref (e.g. "ho_ten" -> "Sheet1!B3").
        output_path   : destination path for the filled xlsx.

        Returns
        -------
        Path to the saved output file.
        """
        template_path = Path(template_path)
        output_path = Path(output_path)

        wb = openpyxl.load_workbook(str(template_path), keep_vba=False)

        for field_id, cell_ref in cell_mapping.items():
            value = form_data.get(field_id)
            m = _SHEET_CELL_RE.match(cell_ref.strip())
            if not m:
                # Skip malformed cell refs silently (caller's responsibility).
                continue
            sheet_name, coord = m.group(1), m.group(2)

            if sheet_name:
                ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            else:
                ws = wb.active

            if ws is not None:
                ws[coord] = value

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))
        return output_path
