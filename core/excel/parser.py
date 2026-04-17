from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from core.models.field_defs import FieldDefinition, FieldType, FieldConstraints


def _slugify(text: str) -> str:
    """Convert a label string to a safe identifier (snake_case, ASCII only)."""
    text = text.lower().strip().rstrip(":")
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = text.strip("_")
    return text or "field"


def _is_bold(cell: object) -> bool:
    """Return True if the openpyxl cell has bold font, safely."""
    font = getattr(cell, "font", None)
    return bool(getattr(font, "bold", False) or False)


def _cell_is_label(cell: object) -> bool:
    """Heuristic: a cell is a label if its value ends with ':' or its font is bold."""
    value = getattr(cell, "value", None)
    if value is None:
        return False
    str_value = str(value).strip()
    if not str_value:
        return False
    if str_value.endswith(":"):
        return True
    if _is_bold(cell):
        return True
    return False


@dataclass
class CellInfo:
    sheet: str
    coord: str
    row: int
    col: int
    value: object
    is_merged: bool
    is_label: bool


@dataclass
class ParseResult:
    cells: list[CellInfo] = field(default_factory=list)
    field_definitions: list[FieldDefinition] = field(default_factory=list)
    raw_text: str = ""


class ExcelParser:
    """Parse an xlsx file and extract cell data plus heuristic FieldDefinitions."""

    def parse(self, path: str | Path) -> ParseResult:
        path = Path(path)
        wb = openpyxl.load_workbook(str(path), data_only=True)

        all_cells: list[CellInfo] = []
        all_field_defs: list[FieldDefinition] = []
        raw_lines: list[str] = []

        for ws in wb.worksheets:
            sheet_name = ws.title
            # Build a set of shadow (non-top-left) merged-cell coordinates.
            merged_shadows: set[tuple[int, int]] = set()
            for merge_range in ws.merged_cells.ranges:
                for row_idx, col_idx in merge_range.cells:
                    if (row_idx, col_idx) != (merge_range.min_row, merge_range.min_col):
                        merged_shadows.add((row_idx, col_idx))

            # Build a coord -> cell map for quick neighbour lookup.
            cell_map: dict[tuple[int, int], object] = {}
            for row in ws.iter_rows():
                for cell in row:
                    cell_map[(cell.row, cell.column)] = cell

            # First pass: collect CellInfo for non-shadow cells.
            sheet_cells: list[CellInfo] = []
            for row in ws.iter_rows():
                for cell in row:
                    rc = (cell.row, cell.column)
                    if rc in merged_shadows:
                        continue  # skip shadow cells
                    is_merged = any(
                        (cell.row, cell.column) == (mr.min_row, mr.min_col)
                        for mr in ws.merged_cells.ranges
                    )
                    is_label = _cell_is_label(cell)
                    ci = CellInfo(
                        sheet=sheet_name,
                        coord=cell.coordinate,
                        row=cell.row,
                        col=cell.column,
                        value=cell.value,
                        is_merged=is_merged,
                        is_label=is_label,
                    )
                    sheet_cells.append(ci)
                    if cell.value is not None:
                        raw_lines.append(f"{sheet_name}!{cell.coordinate}: {cell.value}")

            all_cells.extend(sheet_cells)

            # Second pass: label-to-value pairing → generate FieldDefinitions.
            # Build a lookup by (row, col) → CellInfo for this sheet.
            ci_map: dict[tuple[int, int], CellInfo] = {
                (ci.row, ci.col): ci for ci in sheet_cells
            }

            seen_ids: dict[str, int] = {}

            for ci in sheet_cells:
                if ci.is_label:
                    continue
                if ci.value is None:
                    continue

                # Look left then above for a label cell.
                label_text: str | None = None
                for neighbour_rc in [
                    (ci.row, ci.col - 1),  # left
                    (ci.row - 1, ci.col),  # above
                ]:
                    neighbour = ci_map.get(neighbour_rc)
                    if neighbour and neighbour.is_label and neighbour.value is not None:
                        label_text = str(neighbour.value).strip().rstrip(":")
                        break

                if label_text:
                    base_id = _slugify(label_text)
                    # Deduplicate ids within the parse result.
                    if base_id in seen_ids:
                        seen_ids[base_id] += 1
                        field_id = f"{base_id}_{seen_ids[base_id]}"
                    else:
                        seen_ids[base_id] = 0
                        field_id = base_id

                    fd = FieldDefinition(
                        id=field_id,
                        label=label_text,
                        cell_ref=f"{sheet_name}!{ci.coord}",
                        type=FieldType.TEXT,
                        constraints=FieldConstraints(),
                    )
                    all_field_defs.append(fd)

        return ParseResult(
            cells=all_cells,
            field_definitions=all_field_defs,
            raw_text="\n".join(raw_lines),
        )
